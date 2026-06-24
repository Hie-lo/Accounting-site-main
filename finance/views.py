from multiprocessing import context
from urllib import request

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib import messages
from django.db.models import Sum, Q ,Count
from .models import Transaction, Category, Wallet, recalculate_balance, ManualBalanceHistory , Budget
from .forms import TransactionForm, SignUpForm, BudgetForm
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from .forms import CategoryForm
import json
from datetime import datetime, date ,timedelta
from django.utils import timezone
from calendar import monthrange
from django.db import models
from .utils import jalali_to_gregorian
from jdatetime import date as jdate

def home(request):
    return render(request, 'home.html')

def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'registration/signup.html', {'form': form})

@login_required
def dashboard(request):
    # اطمینان از وجود کیف پول
    wallet, created = Wallet.objects.get_or_create(user=request.user)
    if created:
        default_categories = [
            'خوراک', 'مسکن', 'حمل‌ونقل', 'تفریح', 'بهداشت',
            'تحصیل', 'قبوض', 'خرید روزانه', 'هدیه', 'سایر'
        ]
        for cat_name in default_categories:
            Category.objects.get_or_create(name=cat_name, user=request.user, defaults={'is_default': True})
    
    total_topup = Transaction.objects.filter(user=request.user, type='topup').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = Transaction.objects.filter(user=request.user, type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    balance = wallet.balance
    recent_transactions = Transaction.objects.filter(user=request.user).order_by('-date', '-created_at')[:10]

    # داده‌های نمودار خرج بر اساس دسته
    expense_by_category = (
        Transaction.objects.filter(user=request.user, type='expense')
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    
    # آماده‌سازی برای نمودار
    chart_labels = [item['category__name'] if item['category__name'] else 'بدون دسته' for item in expense_by_category]
    chart_data = [item['total'] for item in expense_by_category]
    
    # اگر هیچ خرجی وجود نداشت، یک پیام پیش‌فرض بدهید
    if not chart_labels:
        chart_labels = ['هیچ داده‌ای']
        chart_data = [0]
    now = datetime.now()
    current_month = now.month
    current_year = now.year
    
    budget_alerts = []
    budgets = Budget.objects.filter(user=request.user, month=current_month, year=current_year)
    for b in budgets:
        spent = Transaction.objects.filter(
            user=request.user,
            type='expense',
            category=b.category,
            date__year=current_year,
            date__month=current_month
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        if b.amount > 0:
            percent = (spent / b.amount) * 100
            if percent >= 90:
                budget_alerts.append(
                    f"⚠️ بودجه دسته «{b.category.name}» به {percent:.0f}% رسیده است. "
                    f"(خرج شده: {spent:,} تومان / بودجه: {b.amount:,} تومان)"
                )

    categories = Category.objects.filter(user=request.user)
    today_date = timezone.now().date().isoformat()
    today = jdate.today()
    context = {
        'total_topup': total_topup,
        'total_expense': total_expense,
        'balance': balance,
        'recent_transactions': recent_transactions,
        'budget_alerts': budget_alerts,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'categories': categories,
        'today': today_date,
        'jalali_year': today.year,
        'jalali_month': today.month,
        'jalali_day': today.day,
    }
 
    return render(request, 'finance/dashboard.html', context)

# finance/views.py
from jdatetime import date as jdate
from .utils import jalali_to_gregorian

@login_required
def add_transaction(request):
    if request.method == 'POST':
        post_data = request.POST.copy()
        
        # دریافت سال، ماه، روز شمسی
        try:
            jy = int(post_data.get('jalali_year'))
            jm = int(post_data.get('jalali_month'))
            jd = int(post_data.get('jalali_day'))
            gregorian_date = jalali_to_gregorian(jy, jm, jd)
            if gregorian_date:
                post_data['date'] = gregorian_date.strftime('%Y-%m-%d')
            else:
                messages.error(request, 'تاریخ وارد شده معتبر نیست.')
                return redirect('transaction_list')
        except (ValueError, TypeError):
            messages.error(request, 'فرمت تاریخ صحیح نیست.')
            return redirect('transaction_list')
        
        form = TransactionForm(post_data, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            recalculate_balance(request.user)
            messages.success(request, 'تراکنش با موفقیت ثبت شد.')
            return redirect('dashboard')
        else:
            messages.error(request, 'اطلاعات وارد شده معتبر نیست.')
            return redirect('transaction_list')
    else:
        # مقدار پیش‌فرض: تاریخ امروز شمسی
        today = jdate.today()
        initial_data = {
            'jalali_year': today.year,
            'jalali_month': today.month,
            'jalali_day': today.day,
        }
        form = TransactionForm(initial=initial_data, user=request.user)
    return render(request, 'finance/add_transaction.html', {'form': form})

@login_required
def transaction_list(request):
    transactions = Transaction.objects.filter(user=request.user).order_by('-date', '-created_at')
    
    # دریافت تاریخ‌های شمسی از GET (اگر قبلاً فیلتر شده باشد)
    from_date_jalali = request.GET.get('from_date')
    to_date_jalali = request.GET.get('to_date')
    
    # ===== تنظیم پیش‌فرض به تاریخ امروز شمسی =====
    today_jalali = jdate.today()
    today_str = today_jalali.strftime('%Y/%m/%d')   # تاریخ امروز برای placeholder
    today_jalali_str = jdate.today().strftime('%Y/%m/%d')
    if today_jalali.month == 12:
        next_month_year = today_jalali.year + 1
        next_month_month = 1
    else:
        next_month_year = today_jalali.year
        next_month_month = today_jalali.month + 1
    
    next_month_jalali = jdate(next_month_year, next_month_month, 1)
    next_month_placeholder = next_month_jalali.strftime('%Y/%m/%d')
    # اگر کاربر تاریخی انتخاب نکرده، امروز را به عنوان پیش‌فرض قرار بده
    from_date_display = from_date_jalali or ''
    to_date_display = to_date_jalali or ''
    # =============================================
    
    # تبدیل تاریخ شمسی به میلادی برای فیلتر کردن (اگر کاربر چیزی وارد کرده باشد)
    if from_date_jalali:
        try:
            jy, jm, jd = map(int, from_date_jalali.split('/'))
            from_date_greg = jalali_to_gregorian(jy, jm, jd)
            transactions = transactions.filter(date__gte=from_date_greg)
        except:
            pass
    if to_date_jalali:
        try:
            jy, jm, jd = map(int, to_date_jalali.split('/'))
            to_date_greg = jalali_to_gregorian(jy, jm, jd)
            transactions = transactions.filter(date__lte=to_date_greg)
        except:
            pass
    
    # فیلتر بر اساس نوع
    type_filter = request.GET.get('type')
    if type_filter and type_filter in ['topup', 'expense']:
        transactions = transactions.filter(type=type_filter)
    
    # فیلتر بر اساس دسته‌بندی
    category_id = request.GET.get('category')
    if category_id:
        transactions = transactions.filter(category_id=category_id)
    
    # جستجو در توضیحات
    search_text = request.GET.get('search')
    if search_text:
        transactions = transactions.filter(Q(description__icontains=search_text))
    
    # فیلتر بر اساس ارزیابی
    rating_filter = request.GET.get('rating')
    if rating_filter and rating_filter in ['bad', 'medium', 'good']:
        transactions = transactions.filter(rating=rating_filter)
    
    # محاسبه تاریخ شمسی برای هر تراکنش (برای دکمه ویرایش)
    for t in transactions:
        if isinstance(t.date, jdate):
            t.jalali_year = t.date.year
            t.jalali_month = t.date.month
            t.jalali_day = t.date.day
        else:
            try:
                jalali_date = jdate.fromgregorian(date=t.date)
                t.jalali_year = jalali_date.year
                t.jalali_month = jalali_date.month
                t.jalali_day = jalali_date.day
            except:
                t.jalali_year = 1400
                t.jalali_month = 1
                t.jalali_day = 1
    
    categories = Category.objects.filter(user=request.user)
    today = jdate.today()
    
    context = {
        'transactions': transactions,
        'categories': categories,
        'type_filter': type_filter,
        'category_id': category_id,
        'from_date': from_date_display,      # مقدار پیش‌فرض امروز شمسی
        'to_date': to_date_display,          # مقدار پیش‌فرض امروز شمسی
        'today_placeholder': today_str,    # برای placeholder «از تاریخ»
        'next_month_placeholder': next_month_placeholder,  # برای placeholder
        'search_text': search_text,
        'rating_filter': rating_filter,
        'jalali_year': today.year,
        'jalali_month': today.month,
        'jalali_day': today.day,
    }
    return render(request, 'finance/transaction_list.html', context)

@login_required
def edit_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    if request.method == 'POST':
        post_data = request.POST.copy()
        try:
            jy = int(post_data.get('jalali_year'))
            jm = int(post_data.get('jalali_month'))
            jd = int(post_data.get('jalali_day'))
            gregorian_date = jalali_to_gregorian(jy, jm, jd)
            if gregorian_date:
                post_data['date'] = gregorian_date.strftime('%Y-%m-%d')
            else:
                messages.error(request, 'تاریخ وارد شده معتبر نیست.')
                return redirect('transaction_list')
        except (ValueError, TypeError):
            messages.error(request, 'فرمت تاریخ صحیح نیست.')
            return redirect('transaction_list')
        
        form = TransactionForm(post_data, instance=transaction, user=request.user)
        if form.is_valid():
            form.save()
            recalculate_balance(request.user)
            messages.success(request, 'تراکنش با موفقیت ویرایش شد.')
            return redirect('transaction_list')
        else:
            messages.error(request, 'خطا در ویرایش تراکنش.')
            return redirect('transaction_list')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    return render(request, 'finance/edit_transaction.html', {'form': form})

@login_required
def delete_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    if request.method == 'POST':
        transaction.delete()
        recalculate_balance(request.user)
        messages.success(request, 'تراکنش حذف شد.')
        return redirect('transaction_list')
    return render(request, 'finance/confirm_delete.html', {'transaction': transaction})


@login_required
def manual_adjust_balance(request):
    wallet = Wallet.objects.get(user=request.user)
    
    # ایجاد یا بازیابی دسته‌بندی ویژه «تنظیم دستی موجودی» برای کاربر
    manual_category, created = Category.objects.get_or_create(
        user=request.user,
        name='تنظیم دستی موجودی',
        defaults={'is_default': False}
    )
    
    if request.method == 'POST':
        new_balance_str = request.POST.get('new_balance')
        reason = request.POST.get('reason')
        if new_balance_str and reason:
            try:
                new_balance = int(new_balance_str)
                old_balance = wallet.balance
                diff = new_balance - old_balance
                
                if diff != 0:
                    # ثبت تراکنش جبرانی با دسته‌بندی مشخص
                    transaction = Transaction.objects.create(
                        user=request.user,
                        amount=abs(diff),
                        type='topup' if diff > 0 else 'expense',
                        category=manual_category,
                        date=timezone.now().date(),
                        description=f"تنظیم دستی موجودی: {reason} (از {old_balance} به {new_balance} تومان)"
                    )
                    recalculate_balance(request.user)
                    ManualBalanceHistory.objects.create(
                        user=request.user,
                        old_balance=old_balance,
                        new_balance=new_balance,
                        reason=reason
                    )
                    messages.success(request, f'موجودی کیف پول از {old_balance:,} به {new_balance:,} تومان تنظیم شد. تراکنش ثبت شد.')
                else:
                    messages.info(request, 'مقدار جدید با موجودی فعلی برابر است.')
            except ValueError:
                messages.error(request, 'مبلغ وارد شده معتبر نیست.')
        else:
            messages.error(request, 'لطفاً مبلغ جدید و دلیل را وارد کنید.')
        return redirect('dashboard')
    
    context = {
        'wallet': wallet,
        'manual_category': manual_category
    }
    return render(request, 'finance/manual_adjust.html', context)

@login_required
def export_excel(request):
    # دریافت فیلترها
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    
    # اعمال فیلترها (مشابه قبل)
    type_filter = request.GET.get('type')
    if type_filter and type_filter in ['topup', 'expense']:
        transactions = transactions.filter(type=type_filter)
    category_id = request.GET.get('category')
    if category_id:
        transactions = transactions.filter(category_id=category_id)
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        transactions = transactions.filter(date__gte=from_date)
    if to_date:
        transactions = transactions.filter(date__lte=to_date)
    search_text = request.GET.get('search')
    if search_text:
        transactions = transactions.filter(description__icontains=search_text)
    
    # ====== فیلتر بر اساس سال و ماه شمسی ======
    year = request.GET.get('year')
    month = request.GET.get('month')
    if year and month:
        try:
            year = int(year)
            month = int(month)
            filtered_transactions = []
            for t in transactions:
                # تبدیل تاریخ به شمسی و مقایسه با سال و ماه ورودی
                if isinstance(t.date, jdate):
                    # تاریخ از نوع jdatetime.date است (شمسی)
                    if t.date.year == year and t.date.month == month:
                        filtered_transactions.append(t)
                else:
                    # تاریخ از نوع datetime.date است (میلادی) - تبدیل به شمسی
                    try:
                        jalali_date = jdate.fromgregorian(date=t.date)
                        if jalali_date.year == year and jalali_date.month == month:
                            filtered_transactions.append(t)
                    except:
                        pass
            transactions = filtered_transactions
        except (ValueError, TypeError):
            pass
    # =========================================
    
    # اگر تراکنشی وجود نداشت، پیام بدهید
    if not transactions:
        messages.warning(request, 'هیچ تراکنشی برای این ماه وجود ندارد.')
        return redirect('monthly_report')
    
    # ایجاد کتاب کار و برگه
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تراکنش‌ها"
    
    # تنظیم جهت راست‌چین برای کل برگه
    ws.sheet_view.rightToLeft = True
    
    # هدرها با استایل زیبا
    headers = ['ردیف', 'نوع', 'مبلغ (تومان)', 'دسته', 'تاریخ', 'ارزیابی', 'توضیحات']
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # رنگ‌بندی برای ارزیابی
    rating_colors = {
        'bad': 'FF0000',    # قرمز
        'medium': 'FFC107', # زرد
        'good': '28A745',   # سبز
    }
    
    # پر کردن داده‌ها
    for row, t in enumerate(transactions, start=2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=t.get_type_display())
        ws.cell(row=row, column=3, value=f"{t.amount:,}")
        ws.cell(row=row, column=4, value=t.category.name if t.category else 'بدون دسته')
        
        # تاریخ به صورت شمسی
        if isinstance(t.date, jdate):
            date_str = t.date.strftime('%Y/%m/%d')
        else:
            try:
                jalali_date = jdate.fromgregorian(date=t.date)
                date_str = jalali_date.strftime('%Y/%m/%d')
            except:
                date_str = str(t.date)
        ws.cell(row=row, column=5, value=date_str)
        
        rating_display = dict(Transaction.RATING_CHOICES).get(t.rating, '')
        rating_cell = ws.cell(row=row, column=6, value=rating_display)
        if t.rating and t.rating in rating_colors:
            rating_cell.fill = openpyxl.styles.PatternFill(
                start_color=rating_colors[t.rating],
                end_color=rating_colors[t.rating],
                fill_type="solid"
            )
            rating_cell.font = Font(color="FFFFFF", bold=True) if t.rating != 'medium' else Font(color="000000", bold=True)
        
        ws.cell(row=row, column=7, value=t.description)
        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    column_widths = [8, 15, 18, 20, 18, 15, 40]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'
    wb.save(response)
    return response


@login_required
def category_list(request):
    categories = Category.objects.filter(user=request.user)
    return render(request, 'finance/category_list.html', {'categories': categories})

@login_required
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            messages.success(request, 'دسته جدید اضافه شد.')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'finance/category_form.html', {'form': form, 'title': 'افزودن دسته'})

@login_required
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'دسته ویرایش شد.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'finance/category_form.html', {'form': form, 'title': 'ویرایش دسته'})

@login_required
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'دسته حذف شد.')
        return redirect('category_list')
    # اگر درخواست GET بود (مثلاً از طریق لینک مستقیم)، خطای 405 بدهید
    from django.http import HttpResponseNotAllowed
    return HttpResponseNotAllowed(['POST'])

@login_required
def monthly_report(request):
    # ========== تعریف ماه‌های شمسی ==========
    persian_months = [
        (1, 'فروردین'), (2, 'اردیبهشت'), (3, 'خرداد'), (4, 'تیر'),
        (5, 'مرداد'), (6, 'شهریور'), (7, 'مهر'), (8, 'آبان'),
        (9, 'آذر'), (10, 'دی'), (11, 'بهمن'), (12, 'اسفند')
    ]
    # =========================================
    
    # دریافت سال و ماه شمسی از GET (پیش‌فرض: ماه جاری شمسی)
    today_jalali = jdate.today()
    selected_jalali_year = int(request.GET.get('year', today_jalali.year))
    selected_jalali_month = int(request.GET.get('month', today_jalali.month))
    
    # تبدیل به میلادی برای فیلتر کردن تراکنش‌ها
    selected_gregorian_date = jalali_to_gregorian(selected_jalali_year, selected_jalali_month, 1)
    if selected_gregorian_date is None:
        # اگر تاریخ نامعتبر بود، از امروز استفاده کن
        selected_gregorian_date = timezone.now().date()
        selected_jalali_year = today_jalali.year
        selected_jalali_month = today_jalali.month
    
    # ساخت بازه تاریخ اول و آخر ماه (میلادی)
    first_day = selected_gregorian_date
    if selected_jalali_month == 12:
        next_month_date = jalali_to_gregorian(selected_jalali_year + 1, 1, 1)
    else:
        next_month_date = jalali_to_gregorian(selected_jalali_year, selected_jalali_month + 1, 1)
    
    if next_month_date:
        last_day = next_month_date - timedelta(days=1)
    else:
        # Fallback: استفاده از monthrange برای تاریخ میلادی
        last_day = date(selected_gregorian_date.year, selected_gregorian_date.month, 
                       monthrange(selected_gregorian_date.year, selected_gregorian_date.month)[1])
    
    # تراکنش‌های آن ماه
    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=first_day,
        date__lte=last_day
    ).order_by('date')
    
    # محاسبه مجموع topup و expense در این ماه
    total_topup = transactions.filter(type='topup').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = transactions.filter(type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    
    # محاسبه موجودی ابتدای ماه (مجموع تراکنش‌های قبل از first_day)
    balance_before = Transaction.objects.filter(
        user=request.user,
        date__lt=first_day
    ).aggregate(
        net=Sum('amount', filter=models.Q(type='topup')) - Sum('amount', filter=models.Q(type='expense'))
    )['net'] or 0
    balance_end = balance_before + total_topup - total_expense
    
    # داده‌های نمودار خرج بر اساس دسته در این ماه
    expense_by_category = (
        transactions.filter(type='expense')
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    chart_labels = [item['category__name'] if item['category__name'] else 'بدون دسته' for item in expense_by_category]
    chart_data = [item['total'] for item in expense_by_category]
    if not chart_labels:
        chart_labels = ['هیچ داده‌ای']
        chart_data = [0]
    
    # داده‌های نمودار افزایش موجودی بر اساس دسته
    topup_by_category = (
        transactions.filter(type='topup')
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    chart_labels_topup = [item['category__name'] if item['category__name'] else 'بدون دسته' for item in topup_by_category]
    chart_data_topup = [item['total'] for item in topup_by_category]
    if not chart_labels_topup:
        chart_labels_topup = ['هیچ داده‌ای']
        chart_data_topup = [0]
    
    # لیست سال‌های شمسی موجود (از اولین تراکنش کاربر تا الان)
    first_transaction = Transaction.objects.filter(user=request.user).order_by('date').first()
    if first_transaction:
        if isinstance(first_transaction.date, jdate):
            min_jalali_year = first_transaction.date.year
        else:
            try:
                min_jalali_date = jdate.fromgregorian(date=first_transaction.date)
                min_jalali_year = min_jalali_date.year
            except:
                min_jalali_year = today_jalali.year
    else:
        min_jalali_year = today_jalali.year
    
    jalali_years = range(min_jalali_year, today_jalali.year + 1)
    
    # نام ماه انتخاب شده
    selected_month_name = dict(persian_months).get(selected_jalali_month, '')
    
    context = {
        'transactions': transactions,
        'total_topup': total_topup,
        'total_expense': total_expense,
        'balance_before': balance_before,
        'balance_end': balance_end,
        'selected_year': selected_jalali_year,
        'selected_month': selected_jalali_month,
        'selected_month_name': selected_month_name,
        'months': persian_months,  # لیست ماه‌های شمسی
        'years': jalali_years,     # لیست سال‌های شمسی
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'chart_labels_topup': json.dumps(chart_labels_topup),
        'chart_data_topup': json.dumps(chart_data_topup),
        # برای لینک اکسل، سال و ماه میلادی را هم ارسال می‌کنیم
        'export_year': selected_gregorian_date.year,
        'export_month': selected_gregorian_date.month,
    }
    return render(request, 'finance/monthly_report.html', context)



@login_required
def budget_list(request):
    budgets = Budget.objects.filter(user=request.user).select_related('category')
    
    # دیکشنری نام ماه‌های شمسی
    persian_months = {
        1: 'فروردین', 2: 'اردیبهشت', 3: 'خرداد', 4: 'تیر',
        5: 'مرداد', 6: 'شهریور', 7: 'مهر', 8: 'آبان',
        9: 'آذر', 10: 'دی', 11: 'بهمن', 12: 'اسفند'
    }
    # دیکشنری نام ماه‌های میلادی (برای داده‌های قدیمی)
    english_months = {
        1: 'ژانویه', 2: 'فوریه', 3: 'مارس', 4: 'آوریل',
        5: 'مه', 6: 'ژوئن', 7: 'ژوئیه', 8: 'اوت',
        9: 'سپتامبر', 10: 'اکتبر', 11: 'نوامبر', 12: 'دسامبر'
    }
    
    # تاریخ امروز شمسی
    today_jalali = jdate.today()
    current_j_month = today_jalali.month
    current_j_year = today_jalali.year
    
    # لیست ماه‌های شمسی برای سلکت
    months_choices = [(num, name) for num, name in persian_months.items()]
    
    # محاسبه spent و percent و اضافه کردن month_name
    for budget in budgets:
        spent = Transaction.objects.filter(
            user=request.user,
            type='expense',
            category=budget.category,
            date__year=budget.year,
            date__month=budget.month
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        budget.spent = spent
        budget.percent = (spent / budget.amount * 100) if budget.amount > 0 else 0
        
        # تشخیص اینکه ماه و سال بودجه میلادی است یا شمسی
        if budget.year > 1300:
            # شمسی
            budget.month_name = persian_months.get(budget.month, '')
        else:
            # میلادی
            budget.month_name = english_months.get(budget.month, '')
    
    categories = Category.objects.filter(user=request.user)
    
    context = {
        'budgets': budgets,
        'current_month': current_j_month,
        'current_year': current_j_year,
        'categories': categories,
        'months': months_choices,
    }
    return render(request, 'finance/budget_list.html', context)

@login_required
def add_budget(request):
    if request.method == 'POST':
        form = BudgetForm(request.POST, user=request.user)
        if form.is_valid():
            budget = form.save(commit=False)
            budget.user = request.user
            budget.save()
            messages.success(request, 'بودجه با موفقیت ثبت شد.')
            return redirect('budget_list')
    else:
        form = BudgetForm(user=request.user)
    return render(request, 'finance/budget_form.html', {'form': form, 'title': 'افزودن بودجه'})

@login_required
def edit_budget(request, pk):
    budget = get_object_or_404(Budget, pk=pk, user=request.user)
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'بودجه ویرایش شد.')
            return redirect('budget_list')
    else:
        form = BudgetForm(instance=budget, user=request.user)
    return render(request, 'finance/budget_form.html', {'form': form, 'title': 'ویرایش بودجه'})

@login_required
def delete_budget(request, pk):
    budget = get_object_or_404(Budget, pk=pk, user=request.user)
    if request.method == 'POST':
        budget.delete()
        messages.success(request, 'بودجه حذف شد.')
        return redirect('budget_list')
    # در صورت درخواست GET، خطای 405
    from django.http import HttpResponseNotAllowed
    return HttpResponseNotAllowed(['POST'])